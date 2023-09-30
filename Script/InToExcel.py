import os
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import win32com.client as win32

column_int = 2

# コマンドプロンプトを非表示に
def hide_console():
    # Windowsのコマンドプロンプトを非表示にする
    if os.name == 'nt':
        try:
            hwnd = win32.GetConsoleWindow()
            win32.ShowWindow(hwnd, 0)
        except Exception:
            pass

hide_console()

def write_text_to_excel(workbook, sheet_name, title, text, delimiter, start_row, start_column):
    # 改行文字("\n")で文章を分割
    if delimiter != "": 
        text = delimiter + delimiter.join(text)
    text_list = text.split("\n")

    # 意図していない空文字の改行分を削除
    while text_list and (not text_list[-1].strip() or text_list[-1] == delimiter):
        text_list.pop()

    # Sheetを選択
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
        
    # IDのタイトル書き込み
    id_title_cell = sheet.cell(row=1, column=1, value='ID')
    
    # IDの値を追加
    for n in range(0,len(text_list)):
        sheet.cell(row=id_title_cell.row + n + start_row, column=1, value=n)

    # タイトルセルを書き込み
    title_cell = sheet.cell(row=start_row, column=start_column, value=title)

    # タイトルセルの下に、順に文章を書き込み
    for i, t in enumerate(text_list):
        sheet.cell(row=title_cell.row+i+1, column=start_column, value=t)


def write_to_excel():
    # Excelのインスタンスを作成 
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    
    # 入力値を取得
    sheet_name = sheet_name_entry.get()
    title = title_entry.get()
    text = text_entry.get("1.0", tk.END)
    delimiter = delimiter_entry.get()
    
    #start_row = int(start_row_entry.get())
    start_row = 1
    start_column = int(start_column_entry.get())
    if(start_row < 1):
        start_row = 1
    if(start_column < 2):
        start_column = 2

    # Excelファイルを開く
    try:
        file_path = file_path_entry.get()
        workbook = openpyxl.load_workbook(file_path)
    except Exception:
        status_label.config(text="Task failed.")

    # Excelに文章を書き込み
    write_text_to_excel(workbook, sheet_name, title, text, delimiter, start_row, start_column)

    # Excelファイルを保存
    workbook.save(file_path)
    status_label.config(text="Task Completed.")

    # Excelファイルを閉じる
    workbook.close()
    # Excelを表示する
    excel.Visible = True
    # 既存のファイルを指定して開く
    workbook = excel.Workbooks.Open(file_path)
    
    
def column_add_value():
    #列行の値加算処理
    global column_int
    column_int += 1
    start_column_var.set(column_int)
    
def column_subtract_value():
    #列行の値加算処理
    global column_int
    if(column_int == 2):
        return
    column_int -= 1
    start_column_var.set(column_int)

# GUIアプリの作成
root = tk.Tk()
root.title("Iton Excel")
root.geometry("400x550")

# 入力項目の作成
file_path_label = ttk.Label(root, text="Excel file path")
file_path_entry = ttk.Entry(root)

sheet_name_label = ttk.Label(root, text="Sheet name")
sheet_name_entry = ttk.Entry(root)

title_label = ttk.Label(root, text="Label")
title_entry = ttk.Entry(root)

text_label = ttk.Label(root, text="Sentence")
text_entry_var = tk.StringVar()
text_entry = tk.Text(root)

delimiter_label = ttk.Label(root, text="Delimiter")
delimiter_entry = ttk.Entry(root)
#エントリーに文字の描画

start_column_label = ttk.Label(root, text="開始列番号")
start_column_var = tk.StringVar()
start_column_var.set(column_int)
start_column_entry = ttk.Entry(root, textvariable=start_column_var)

#処理のログ
status_label = tk.Label(root, text="")

# ボタンの作成
file_select_button = ttk.Button(root, text="Excelファイルを選択", command=lambda: file_path_entry.insert(tk.END, filedialog.askopenfilename()))
write_button = ttk.Button(root, text="Excelに書き込む", command=write_to_excel)
column_set_add_button = ttk.Button(root, text="+", command=column_add_value)
column_set_subtract_button = ttk.Button(root, text="-", command=column_subtract_value)

# 入力項目とボタンの配置
file_path_label.place(x=75, y=25)
file_select_button.place(x=60, y=50, width=110, height=25)
file_path_entry.place(x=180, y=50, width=170, height=25)

sheet_name_label.place(x=80, y=105)
sheet_name_entry.place(x=180, y=100, width=170, height=25)

title_label.place(x=95, y=155)
title_entry.place(x=180, y=150, width=170, height=25)

text_label.place(x=88, y=205)
text_entry.place(x=180, y=200, width=170, height=200)

delimiter_label.place(x=95, y=425)
delimiter_entry.place(x=180, y=425, width=20, height=20)

start_column_label.place(x=90, y=465)
column_set_subtract_button.place(x=180, y=465, width=20, height=20)
start_column_entry.place(x=240, y=465, width=20, height=20)
column_set_add_button.place(x=300, y=465, width=20, height=20)
write_button.place(x=150, y=505)

status_label.place(x=280, y=505)

root.mainloop()